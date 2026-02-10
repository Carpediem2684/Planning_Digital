import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Pas de st.set_page_config ici : il est déjà dans app.py


def show_dashboard_pic():
    
    # === Bouton retour menu ===
    st.markdown("""
        <style>
        .back-btn {
            background: #1F3C88;
            color: white;
            padding: 10px 22px;
            border-radius: 8px;
            font-weight: 700;
            cursor:pointer;
            transition: 0.2s;
        }
        .back-btn:hover {
            background: #2750C4;
        }
        </style>
    """, unsafe_allow_html=True)

    col_back, _ = st.columns([1, 5])
    with col_back:
        if st.button("⬅️ Retour Menu"):
            st.session_state["page"] = "menu"
            return
    # === Paramètres GIF ===
    GIF_PATH = 'GIF_20251219_081101_562.gif'  # chemin local

    # Chargement des données
    file_path = 'Essai appli dashboard (1).xlsx'
    df = pd.read_excel(file_path, sheet_name='2025', engine='openpyxl', header=None)

    # --- Lecture de la cellule Q4 pour l’En-cours Visitage ---
    wb_info = load_workbook(file_path, data_only=True)
    ws_info = wb_info["2025"]  # ou le nom exact si différent

    en_cours_visitage = ws_info["Q4"].value

    if en_cours_visitage is None:
       en_cours_visitage = 0

    # === Soucis de cylindre via openpyxl : lecture robuste d'AJ à AM ===
    try:
        wb = load_workbook(file_path, data_only=True)

        # On cible la feuille "2025" si elle existe, sinon on prend la 1ère
        sheet_name_cyl = '2025'
        if sheet_name_cyl not in wb.sheetnames:
            sheet_name_cyl = wb.sheetnames[0]

        ws = wb[sheet_name_cyl]

        # Colonnes Excel -> index openpyxl (A=1) : AJ=36, AK=37, AL=38, AM=39
        COL_AJ, COL_AK, COL_AL, COL_AM = 36, 37, 38, 39

        rows_list = []
        # On parcourt de la ligne 2 jusqu'à la fin de la feuille
        for r in range(2, ws.max_row + 1):
            cylindre = ws.cell(row=r, column=COL_AJ).value  # AJ
            delai    = ws.cell(row=r, column=COL_AK).value  # AK
            retour   = ws.cell(row=r, column=COL_AL).value  # AL
            impact   = ws.cell(row=r, column=COL_AM).value  # AM

            # Ignore les lignes où le cylindre n'est pas renseigné
            if cylindre is None or (isinstance(cylindre, str) and cylindre.strip() == ""):
                continue

            rows_list.append({
                "Cylindre": cylindre,
                "Délai": delai,
                "Retour prévu": retour,
                "Impact client": impact
            })

        # Conversion en DataFrame pandas
        issues = pd.DataFrame(rows_list)

        if not issues.empty:
            # Normalise la date (AL)
            issues["Retour prévu"] = pd.to_datetime(
                issues["Retour prévu"], errors="coerce", dayfirst=True
            )
            # Colonne pour affichage
            issues["Retour prévu (aff.)"] = issues["Retour prévu"].dt.strftime("%d/%m/%Y")
            # Tri par date de retour (les NaT passent en bas)
            issues = issues.sort_values(by=["Retour prévu"], na_position="last").reset_index(drop=True)

            # KPIs
            nb_cylindres = len(issues)
            prochaine_date = issues["Retour prévu"].dropna().min()
            prochaine_date_aff = prochaine_date.strftime("%d/%m/%Y") if pd.notna(prochaine_date) else "—"
        else:
            nb_cylindres = 0
            prochaine_date_aff = "—"

    except Exception as e:
        # Repli propre si jamais la lecture échoue
        issues = pd.DataFrame(columns=["Cylindre", "Délai", "Retour prévu (aff.)", "Impact client"])
        nb_cylindres = 0
        prochaine_date_aff = "—"
        st.warning(f"Impossible de lire AJ:AM (soucis de cylindre). Détail : {e}")

    # === Chargement du calendrier des postes ===
    calendrier_path = 'Calendrier 2026.xlsx'
    df_cal = pd.read_excel(calendrier_path, sheet_name='Feuil1', engine='openpyxl')

    # Renommer proprement les colonnes du calendrier
    df_cal.columns = [
        'Jour',  # Date
        'Horaire_1', 'Etat_1',
        'Horaire_2', 'Etat_2',
        'Horaire_3', 'Etat_3'
    ]

    # Conversion de la colonne date au bon format
    df_cal['Jour'] = pd.to_datetime(df_cal['Jour'], dayfirst=True)

    # Calcul du nombre de postes ouverts par jour (0 à 3)
    df_cal['Postes_ouverts'] = (
        (df_cal['Etat_1'] == 'OUVERT').astype(int) +
        (df_cal['Etat_2'] == 'OUVERT').astype(int) +
        (df_cal['Etat_3'] == 'OUVERT').astype(int)
    )

    # Initialisation
    mois = df.iloc[2:14, 0].tolist()
    campagnes = df.iloc[1, 25:34].tolist()  # Colonnes Z à AH incluses

    pic_realise = pd.Series(
        pd.to_numeric(df.iloc[2:14, 1], errors='coerce').fillna(0).astype(int).values,
        index=mois
    )
    pic_prevu = pd.Series(
        pd.to_numeric(df.iloc[2:14, 2], errors='coerce').fillna(0).astype(int).values,
        index=mois
    )
    ruptures = int(df.iloc[1, 16])

    # Taux d'adhérence global (W2)
    raw_adherence = pd.to_numeric(df.iloc[1, 22], errors='coerce')
    taux_adherence = (raw_adherence * 100) if pd.notna(raw_adherence) else 0

    # Taux d'adhérence S-1 (T2)
    adherence_s1 = pd.to_numeric(df.iloc[1, 19], errors='coerce')

    # Sidebar
    st.sidebar.image(
        "https://upload.wikimedia.org/wikipedia/commons/thumb/3/3f/Logo_Gerflor.svg/2560px-Logo_Gerflor.svg.png",
        width=150
    )
    st.sidebar.title("Sélection UAP")
    uap_selection = st.sidebar.selectbox("Choisir une UAP", ["4M", "2M", "P2000", "KLAM"])
    mois_selectionne = st.sidebar.selectbox("Choisir un mois", mois)

    # Données campagnes (Z à AH)
    campagne_data = df.iloc[2:14, 25:34]
    campagne_data.columns = campagnes
    campagne_data.index = mois
    campagne_mois = campagne_data.loc[mois_selectionne].apply(pd.to_numeric, errors='coerce').fillna(0)

    # Données hebdomadaires
    weekly_data = df.iloc[2:51, [21, 22]]
    weekly_data.columns = ["Semaine", "Taux d'adhérence"]
    weekly_data.dropna(inplace=True)
    weekly_data["Taux d'adhérence"] = pd.to_numeric(
        weekly_data["Taux d'adhérence"], errors="coerce"
    )
    weekly_data["Taux d'adhérence"] = (weekly_data["Taux d'adhérence"] * 100).round(1)
    weekly_data["Semaine"] = weekly_data["Semaine"].astype(int)
    semaines_completes = list(range(1, 51))
    colors = ["green" if val >= 85 else "red" for val in weekly_data["Taux d'adhérence"]]

    # --- État session pour le bouton Félicitations ---
    if "gif_visible" not in st.session_state:
        st.session_state.gif_visible = False
    if "mois_selectionne" not in st.session_state:
        st.session_state.mois_selectionne = mois_selectionne

    if "current_value" not in st.session_state or st.session_state.mois_selectionne != mois_selectionne:
        # Reset de l'état au changement de mois
        st.session_state.current_value = pic_realise[mois_selectionne]
        st.session_state.campagne_clicks = {campagne: False for campagne in campagnes}
        st.session_state.mois_selectionne = mois_selectionne
        st.session_state.bar_color = "darkblue"
        # On masque le GIF au changement de mois
        st.session_state.gif_visible = False

    # Définir les couleurs pour chaque campagne
    couleurs_campagnes = {
        campagnes[0]: "green",
        campagnes[1]: "purple",
        campagnes[2]: "orange",
        campagnes[3]: "pink",
        campagnes[4]: "cyan",
        campagnes[5]: "brown",
        campagnes[6]: "blue",
        campagnes[7]: "magenta",
        campagnes[8]: "lime"
    }

    # Titre et date
    st.markdown(
        f"<h1 style='text-align:center; color:#ffffff;'>Dashboard PIC - {uap_selection}</h1>",
        unsafe_allow_html=True
    )
    date_du_jour = datetime.today().strftime('%d/%m/%Y')
    st.markdown(
        f"<p style='text-align:right; font-size:16px; font-weight:bold;'>Date du jour : {date_du_jour}</p>",
        unsafe_allow_html=True
    )
    # === Encadré "Soucis de cylindre" – haut à droite ===
    left_spacer, right_panel = st.columns([3, 2])  # Ajuste le ratio si besoin

    with right_panel:
        st.markdown(
            """
            <div style="
                background: linear-gradient(135deg, #0F1730 0%, #1E2B57 100%);
                border: 1px solid #23315f;
                border-radius: 12px; padding: 14px 16px; color: #ffffff;">
                <div style="display:flex; align-items:center; justify-content:space-between;">
                    <div style="font-weight:700; font-size:18px;">
                        ⚠️ Cylindres hors service
                    </div>
                    <div style="
                        background:#FFB200; color:#1b1b1b; font-weight:700;
                        padding:4px 10px; border-radius:999px; font-size:12px;">
                        {badge}
                    </div>
                </div>
                <div style="margin-top:8px; font-size:13px; opacity:0.9;">
                    Prochaine date de retour prévue : <b>{next_back}</b>
                </div>
            </div>
            """.format(
                badge=f"{nb_cylindres} en cours" if nb_cylindres else "Aucun",
                next_back=prochaine_date_aff
            ),
            unsafe_allow_html=True
        )

        if nb_cylindres:
            # Tableau compact pour lecture rapide
            cols_aff = ["Cylindre", "Délai", "Retour prévu (aff.)", "Impact client"]
            st.dataframe(
                issues[cols_aff].rename(columns={"Retour prévu (aff.)": "Retour prévu"}),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("Aucun souci de cylindre en cours ✅")

    # --- Bouton Félicitations (au-dessus des métriques) ---
    col_btn = st.container()
    with col_btn:
        if pic_realise[mois_selectionne] > pic_prevu[mois_selectionne]:
            # Libellé dynamique
            label = "🎉 Félicitations (afficher le GIF)" if not st.session_state.gif_visible else "❌ Masquer le GIF"
            if st.button(label):
                st.session_state.gif_visible = not st.session_state.gif_visible
        else:
            # Rien n'apparaît si pas de dépassement
            pass

    # Affichage du GIF en grand si gif_visible == True
    if st.session_state.gif_visible:
        st.markdown("<div style='text-align:center;'>", unsafe_allow_html=True)
        st.image(GIF_PATH, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # --- Section Suivi Objectif Journalier ---
    st.markdown("### 📊 Suivi Objectif Journalier")

    # Données PIC
    pic_total = pic_prevu[mois_selectionne]
    pic_realise_val = pic_realise[mois_selectionne]
    pic_restant = pic_total - pic_realise_val

    # Normalisation de la date du jour (pour inclure la journée en cours)
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    # Détermination de la fin de mois
    if today.month < 12:
        next_month = datetime(today.year, today.month + 1, 1)
    else:
        next_month = datetime(today.year + 1, 1, 1)
    end_date = next_month - timedelta(days=1)

    # Trouver le dernier vendredi du mois
    last_friday = end_date
    while last_friday.weekday() != 4:  # 4 = vendredi
        last_friday -= timedelta(days=1)

    # --- Utilisation du calendrier réel pour calculer jours/postes restants ---
    masque_periode = (df_cal['Jour'] >= today) & (df_cal['Jour'] <= last_friday)
    df_cal_periode = df_cal[masque_periode]

    # Jours restants = nb de jours avec au moins un poste ouvert
    jours_restants = df_cal_periode.loc[df_cal_periode['Postes_ouverts'] > 0, 'Jour'].nunique()

    # Postes restants = somme des postes ouverts sur la période
    postes_restants = df_cal_periode['Postes_ouverts'].sum()

    # Calculs des objectifs
    objectif_par_poste = pic_restant / postes_restants if postes_restants > 0 else 0
    objectif_journalier = pic_restant / jours_restants if jours_restants > 0 else 0

    # --- Affichage ---
    st.markdown("### 📊 Objectifs recalculés (dynamiques)")
    st.markdown(f"""
    - **PIC prévu** : {pic_total} km²  
    - **PIC réalisé** : {pic_realise_val} km²  
    - **PIC restant** : {pic_restant} km²  
    - **Jours restants (avec au moins un poste ouvert)** : {jours_restants}  
    - **Postes restants (réels)** : {postes_restants}  
    - **Objectif par poste** : {objectif_par_poste:.1f} km²  
    - **Objectif journalier moyen** : {objectif_journalier:.1f} km²  
    """)

    # Affichage des métriques
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("PIC Réalisé", f"{pic_realise[mois_selectionne]} km²")
    col1b, col2b, col3b, col4b = st.columns(4)
    col1b.metric("En-cours Visitage", f"{en_cours_visitage} km²")
    col2.metric("PIC Prévu", f"{pic_prevu[mois_selectionne]} km²")
    col3.metric("Ruptures cette semaine", f"{ruptures}")
    col4.metric("Taux d'adhérence S-1", f"{adherence_s1:.1f}%" if pd.notna(adherence_s1) else "N/A")
   

    # Graphiques côte à côte
    campagne_labels = df.iloc[1, 6:14].tolist()
    campagne_values = df[df.iloc[:, 0] == mois_selectionne].iloc[0, 6:14]
    campagne_values = pd.to_numeric(campagne_values, errors='coerce').fillna(0)

    couleurs_personnalisees = {
        "PRIMETEX": "yellow", "TEXLINE": "green", "NERA": "blue", "MOUSSE": "red",
        "TARABUS": "lightgreen", "SPORISOL": "lightgrey", "START": "grey", "TMAX": "brown"
    }
    colors_pie = [couleurs_personnalisees.get(label, "white") for label in campagne_labels]

    fig_pie = go.Figure(data=[
        go.Pie(
            labels=campagne_labels,
            values=campagne_values,
            hole=0.4,
            textinfo='label+percent+value',
            marker=dict(colors=colors_pie)
        )
    ])
    fig_pie.update_layout(title="Répartition par campagne", height=400)

    fig_weekly = go.Figure()
    fig_weekly.add_trace(go.Scatter(
        x=weekly_data["Semaine"],
        y=weekly_data["Taux d'adhérence"],
        mode='markers+lines+text',
        marker=dict(color=colors, size=10),
        name="Taux d'adhérence",
        text=[f"{val:.1f}%" for val in weekly_data["Taux d'adhérence"]],
        textposition="top center"
    ))
    fig_weekly.add_trace(go.Scatter(
        x=semaines_completes,
        y=[85] * len(semaines_completes),
        mode='lines',
        name="Objectif",
        line=dict(dash='dash', color='blue')
    ))
    fig_weekly.update_layout(
        title="Évolution hebdomadaire du taux d'adhérence",
        height=400,
        xaxis_title="Semaine",
        yaxis_title="% d'adhérence"
    )

    col_pie, col_weekly = st.columns(2)
    with col_pie:
        st.plotly_chart(fig_pie, use_container_width=True)
    with col_weekly:
        st.plotly_chart(fig_weekly, use_container_width=True)

    st.markdown(
        "<div style='padding:10px; background-color:#ffdddd; border-left:5px solid red;'>"
        "<p style='color:#a00000; font-size:20px; font-weight:bold; text-align:center;'>"
        "⚠️ Section en phase de test : certaines données peuvent être inexactes."
        "</p></div>",
        unsafe_allow_html=True
    )

    # --- Section Campagnes restantes du mois ---
    st.markdown("### Campagnes restantes du mois")
    cols = st.columns(len(campagnes) + 1)

    # Bouton reset
    if cols[0].button("🔄 Instant T"):
        st.session_state.current_value = pic_realise[mois_selectionne]
        st.session_state.campagne_clicks = {campagne: False for campagne in campagnes}
        st.session_state.bar_color = "darkblue"
        st.session_state.adjustments = {campagne: 0.0 for campagne in campagnes}

    # Initialisation si non présents
    if "adjustments" not in st.session_state:
        st.session_state.adjustments = {campagne: 0.0 for campagne in campagnes}
    if "campagne_clicks" not in st.session_state:
        st.session_state.campagne_clicks = {campagne: False for campagne in campagnes}

    # Affichage des boutons + champs d'ajustement
    for i, campagne in enumerate(campagnes):
        val = campagne_mois[campagne]
        if val > 0:
            clicked = st.session_state.campagne_clicks[campagne]
            indicator = "🟢" if not clicked else "🔴"

            # Champ ajustement
            adj = cols[i + 1].number_input(
                f"Ajustement {campagne} (km²)",
                value=float(st.session_state.adjustments[campagne]),
                step=10.0,
                format="%.1f"
            )
            st.session_state.adjustments[campagne] = adj

            # Bouton toggle
            if cols[i + 1].button(f"{indicator} {campagne}"):
                if not clicked:
                    # Ajout
                    st.session_state.campagne_clicks[campagne] = True
                    st.session_state.current_value += val + adj
                    st.session_state.bar_color = couleurs_campagnes.get(campagne, "darkblue")
                else:
                    # Retrait
                    st.session_state.campagne_clicks[campagne] = False
                    st.session_state.current_value -= val + adj
                    st.session_state.bar_color = "darkblue"

    # ✅ Jauge dynamique
    fig_dynamic = go.Figure(go.Indicator(
        mode="gauge+number",
        value=st.session_state.current_value,
        title={'text': f"Progression PIC ({mois_selectionne})"},
        gauge={
            'axis': {'range': [0, pic_prevu[mois_selectionne] * 1.2]},
            'bar': {'color': st.session_state.bar_color},
            'steps': [
                {'range': [0, pic_prevu[mois_selectionne] * 0.85], 'color': "lightgreen"},
                {'range': [pic_prevu[mois_selectionne] * 0.85, pic_prevu[mois_selectionne]], 'color': "yellow"},
                {'range': [pic_prevu[mois_selectionne], pic_prevu[mois_selectionne] * 1.2], 'color': "lightgrey"}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': pic_prevu[mois_selectionne]
            }
        }
    ))
    st.plotly_chart(fig_dynamic, use_container_width=True)

    # ✅ Message dépassement
    if st.session_state.current_value > pic_prevu[mois_selectionne]:
        st.markdown(
            f"<p style='color:red; font-size:18px; font-weight:bold;'>⚠ Dépassement du PIC prévu : {st.session_state.current_value} km²</p>",
            unsafe_allow_html=True
        )

    # ✅ Tableau des ajustements
    st.markdown("#### Ajustements appliqués")
    st.write(pd.DataFrame.from_dict(st.session_state.adjustments, orient='index', columns=['Ajustement (km²)']))

    # Heatmap
    campagne_data_heatmap = df.iloc[2:14, 6:14]
    campagne_data_heatmap.columns = campagne_labels
    campagne_data_heatmap.index = mois
    campagne_data_heatmap = campagne_data_heatmap.apply(pd.to_numeric, errors='coerce').fillna(0)

    fig_heatmap = go.Figure(data=go.Heatmap(
        z=campagne_data_heatmap.values,
        x=campagne_data_heatmap.columns,
        y=campagne_data_heatmap.index,
        colorscale='Viridis',
        colorbar=dict(title="Valeur"),
        zmin=0,
        zmax=campagne_data_heatmap.values.max(),
        hoverongaps=False
    ))

    annotations = []
    for i, mois_val in enumerate(campagne_data_heatmap.index):
        for j, campagne_val in enumerate(campagne_data_heatmap.columns):
            value = campagne_data_heatmap.iloc[i, j]
            annotations.append(dict(
                x=campagne_val,
                y=mois_val,
                text=str(value),
                showarrow=False,
                font=dict(
                    color="white" if value < campagne_data_heatmap.values.max() / 2 else "black",
                    size=10
                )
            ))

    fig_heatmap.update_layout(
        title="Heatmap des campagnes (améliorée)",
        height=600,
        annotations=annotations,
        xaxis=dict(title="Campagnes", tickangle=-45),
        yaxis=dict(title="Mois")
    )

    st.plotly_chart(fig_heatmap, use_container_width=True)

